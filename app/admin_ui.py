from __future__ import annotations
import csv
import io
from typing import Annotated
from fastapi import Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.routing import APIRouter
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, joinedload
from app.config import settings
from app.db import get_db
from app.models import Classroom, Enrollment, Node, Student, now_utc
from app.attestation import write_audit
from app.sms import send_sms

router = APIRouter(prefix="/admin", tags=["admin-ui"])
templates = Jinja2Templates(directory="app/templates")


def _key_ok(key: str) -> bool:
    return bool(key) and key == settings.token_secret


# ---------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------

@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
def admin_home(request: Request, key: str = "", msg: str = ""):
    if not key or not _key_ok(key):
        return templates.TemplateResponse("admin_login.html",
            {"request": request, "error": "Λάθος κλειδί." if (key and not _key_ok(key)) else ""})
    db = next(get_db())
    nodes = db.query(Node).options(
        joinedload(Node.classrooms).joinedload(Classroom.enrollments),
    ).order_by(Node.id).all()
    db.close()
    return templates.TemplateResponse("admin_home.html",
        {"request": request, "key": key, "nodes": nodes, "msg": msg})


@router.post("", response_class=HTMLResponse)
def admin_login(request: Request, key: str = Form("")):
    if not _key_ok(key):
        return templates.TemplateResponse("admin_login.html",
            {"request": request, "error": "Λάθος κλειδί Admin."})
    return RedirectResponse(f"/admin?key={key}", status_code=303)


# ---------------------------------------------------------------------------
# Node management (admin-only)
# ---------------------------------------------------------------------------

@router.post("/nodes/create")
def create_node(
    db: Annotated[Session, Depends(get_db)],
    key: str = Form(""),
    name: str = Form(...),
    municipality: str = Form("ΘΕΣΣΑΛΟΝΙΚΗΣ"),
    responsible_name: str = Form(""),
    capacity: int = Form(15),
):
    if not _key_ok(key):
        return RedirectResponse("/admin", status_code=303)
    existing = db.query(Node).filter(Node.name == name.strip()).first()
    if existing:
        return RedirectResponse(f"/admin?key={key}&msg=Ο+κόμβος+υπάρχει+ήδη", status_code=303)
    node = Node(
        name=name.strip(),
        municipality=municipality.strip() or "ΘΕΣΣΑΛΟΝΙΚΗΣ",
        responsible_name=responsible_name.strip(),
        capacity=capacity,
    )
    db.add(node)
    db.commit()
    write_audit(db, "admin_node_created", "node", node.id,
                actor="admin", detail={"name": node.name})
    return RedirectResponse(f"/admin?key={key}&msg=Κόμβος+δημιουργήθηκε", status_code=303)


# ---------------------------------------------------------------------------
# CSV / Excel import
# ---------------------------------------------------------------------------

@router.get("/import", response_class=HTMLResponse)
def import_page(request: Request, key: str = ""):
    if not _key_ok(key):
        return RedirectResponse("/admin", status_code=303)
    db = next(get_db())
    nodes = db.query(Node).order_by(Node.name).all()
    db.close()
    return templates.TemplateResponse("admin_import_csv.html",
        {"request": request, "key": key, "nodes": nodes, "result": None})


@router.post("/import/upload")
async def import_upload(
    request: Request,
    db: Annotated[Session, Depends(get_db)],
    key: str = Form(""),
    node_id: str = Form(""),
    new_node_name: str = Form(""),
    municipality: str = Form("ΘΕΣΣΑΛΟΝΙΚΗΣ"),
    classroom_name: str = Form(...),
    teacher_name: str = Form(""),
    teacher_phone: str = Form(""),
    target_hours: int = Form(40),
    file: UploadFile = File(...),
):
    if not _key_ok(key):
        return RedirectResponse("/admin", status_code=303)

    # Resolve or create node
    if node_id and node_id.isdigit():
        node = db.query(Node).filter(Node.id == int(node_id)).first()
    else:
        node = None

    if not node and new_node_name.strip():
        node = db.query(Node).filter(Node.name == new_node_name.strip()).first()
        if not node:
            node = Node(
                name=new_node_name.strip(),
                municipality=municipality.strip() or "ΘΕΣΣΑΛΟΝΙΚΗΣ",
                responsible_name=teacher_name.strip(),
                capacity=30,
            )
            db.add(node)
            db.flush()

    if not node:
        nodes = db.query(Node).order_by(Node.name).all()
        return templates.TemplateResponse("admin_import_csv.html", {
            "request": request, "key": key, "nodes": nodes,
            "result": {"error": "Επιλέξτε ή δημιουργήστε κόμβο."},
        })

    # Parse file
    content = await file.read()
    filename = (file.filename or "").lower()
    rows = []

    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")  # handles BOM from Excel exports
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or "").strip() for v in row])))
        else:
            nodes = db.query(Node).order_by(Node.name).all()
            return templates.TemplateResponse("admin_import_csv.html", {
                "request": request, "key": key, "nodes": nodes,
                "result": {"error": "Αποδεκτοί μόνο .csv και .xlsx"},
            })
    except Exception as exc:
        nodes = db.query(Node).order_by(Node.name).all()
        return templates.TemplateResponse("admin_import_csv.html", {
            "request": request, "key": key, "nodes": nodes,
            "result": {"error": f"Σφάλμα ανάγνωσης αρχείου: {exc}"},
        })

    # Column name aliases
    def _get(row: dict, *keys: str) -> str:
        for k in keys:
            if k in row and row[k]:
                return str(row[k]).strip()
        return ""

    # Create classroom
    classroom = Classroom(
        node_id=node.id,
        name=classroom_name.strip(),
        teacher_name=teacher_name.strip(),
        teacher_phone=teacher_phone.strip(),
        location=node.name,
        target_teaching_hours=target_hours,
        l2e_tenant_id=settings.l2e_tenant_id,
    )
    db.add(classroom)
    db.flush()

    created = []
    skipped = []
    for i, row in enumerate(rows, 1):
        full_name = _get(row, "full_name", "ονοματεπώνυμο", "name", "όνομα")
        if not full_name:
            skipped.append(f"Γραμμή {i}: κενό όνομα")
            continue
        phone = _get(row, "phone", "τηλέφωνο", "mobile", "tel")
        ext_ref = _get(row, "external_ref", "ref", "κωδικός", "id")
        gender = _get(row, "gender", "φύλο")
        student = Student(
            node_id=node.id,
            full_name=full_name,
            phone=phone,
            external_ref=ext_ref or f"IMP-{i:03d}",
            gender=gender,
            status="selected",
            priority_order=i,
        )
        db.add(student)
        db.flush()
        db.add(Enrollment(classroom_id=classroom.id, student_id=student.id, status="active"))
        created.append(full_name)

    db.commit()
    write_audit(db, "admin_csv_import", "classroom", classroom.id,
                actor="admin", detail={"created": len(created), "skipped": len(skipped)})

    return templates.TemplateResponse("admin_import_csv.html", {
        "request": request, "key": key,
        "nodes": db.query(Node).order_by(Node.name).all(),
        "result": {
            "ok": True,
            "classroom_id": classroom.id,
            "classroom_name": classroom.name,
            "node_name": node.name,
            "created": created,
            "skipped": skipped,
        },
    })


# ---------------------------------------------------------------------------
# Phone management per classroom
# ---------------------------------------------------------------------------

@router.get("/class/{class_id}", response_class=HTMLResponse)
def admin_class(
    request: Request, class_id: int,
    key: str = "", sms_sent: int = 0, sms_failed: int = 0, saved: int = 0,
):
    if not _key_ok(key):
        return RedirectResponse("/admin", status_code=303)
    db = next(get_db())
    classroom = db.query(Classroom).options(
        joinedload(Classroom.node),
        joinedload(Classroom.enrollments).joinedload(Enrollment.student),
    ).filter(Classroom.id == class_id).first()
    db.close()
    if not classroom:
        return RedirectResponse(f"/admin?key={key}", status_code=303)
    enrollments = sorted(
        [e for e in classroom.enrollments if e.status == "active"],
        key=lambda e: e.student.full_name or "",
    )
    return templates.TemplateResponse("admin_phones.html", {
        "request": request, "key": key, "classroom": classroom,
        "enrollments": enrollments,
        "sms_sent": sms_sent, "sms_failed": sms_failed, "saved": saved,
    })


@router.post("/class/{class_id}/update-phones")
async def update_phones(
    request: Request, class_id: int,
    db: Annotated[Session, Depends(get_db)],
    key: str = Form(""),
):
    if not _key_ok(key):
        return RedirectResponse("/admin", status_code=303)
    form = await request.form()
    saved = 0
    for field, value in form.items():
        if field.startswith("phone_"):
            student_id = int(field.split("_", 1)[1])
            student = db.query(Student).filter(Student.id == student_id).first()
            if student:
                student.phone = str(value).strip()
                saved += 1
    db.commit()
    write_audit(db, "admin_phones_updated", "classroom", class_id,
                actor="admin", detail={"updated": saved})
    return RedirectResponse(f"/admin/class/{class_id}?key={key}&saved={saved}", status_code=303)


@router.post("/class/{class_id}/send-sms")
def admin_send_sms(
    class_id: int,
    db: Annotated[Session, Depends(get_db)],
    key: str = Form(""),
    message_template: str = Form("Γεια {name}! Παρακολουθήστε την παρουσία σας στο ThrEDU: {url}"),
):
    if not _key_ok(key):
        return RedirectResponse("/admin", status_code=303)
    classroom = db.query(Classroom).options(
        joinedload(Classroom.enrollments).joinedload(Enrollment.student),
    ).filter(Classroom.id == class_id).first()
    if not classroom:
        return RedirectResponse(f"/admin?key={key}", status_code=303)
    enrollments = [e for e in classroom.enrollments if e.status == "active"]
    sent = failed = skipped = 0
    for enr in enrollments:
        s = enr.student
        if not s or not s.phone:
            skipped += 1
            continue
        first_name = s.full_name.split()[0] if s.full_name else s.full_name
        url = f"{settings.public_base_url}/classes/{class_id}"
        body = message_template.format(name=first_name, url=url)
        result = send_sms(s.phone, body)
        if result.ok:
            sent += 1
        else:
            failed += 1
    write_audit(db, "admin_bulk_sms", "classroom", class_id,
                actor="admin", detail={"sent": sent, "failed": failed, "skipped": skipped})
    db.commit()
    return RedirectResponse(
        f"/admin/class/{class_id}?key={key}&sms_sent={sent}&sms_failed={failed}",
        status_code=303,
    )
