"""CSV import and Excel export functionality."""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Tuple
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models import Student, Classroom, Enrollment, Attendance, Lesson


def import_students_to_classroom(
    classroom: Classroom,
    csv_content: str,
    db: Session,
    skip_header: bool = True
) -> tuple[int, int, list[str]]:
    """
    Import students from CSV content into an existing classroom.

    CSV format (columns):
    - full_name (required)
    - phone (optional)
    - email (optional)
    - external_ref (optional)
    - tax_id (optional)
    - gender (optional)

    Returns: (imported_count, skipped_count, error_messages)
    """
    imported = skipped = 0
    errors = []

    try:
        reader = csv.DictReader(
            io.StringIO(csv_content),
            fieldnames=['full_name', 'phone', 'email', 'external_ref', 'tax_id', 'gender']
        )

        for row_num, row in enumerate(reader, start=1):
            if skip_header and row_num == 1:
                continue

            full_name = (row.get('full_name') or '').strip()
            if not full_name:
                errors.append(f"Row {row_num}: Missing full_name")
                skipped += 1
                continue

            phone = (row.get('phone') or '').strip()
            email = (row.get('email') or '').strip()
            external_ref = (row.get('external_ref') or '').strip()
            tax_id = (row.get('tax_id') or '').strip()
            gender = (row.get('gender') or '').strip()

            existing = db.query(Student).filter(
                Student.full_name == full_name,
                Student.node_id == classroom.node_id
            ).first()

            if existing:
                student = existing
            else:
                student = Student(
                    node_id=classroom.node_id,
                    full_name=full_name,
                    phone=phone,
                    email=email,
                    external_ref=external_ref,
                    tax_id=tax_id,
                    gender=gender,
                    status="selected"
                )
                db.add(student)
                db.flush()

            existing_enrollment = db.query(Enrollment).filter(
                Enrollment.classroom_id == classroom.id,
                Enrollment.student_id == student.id
            ).first()

            if not existing_enrollment:
                enrollment = Enrollment(
                    classroom_id=classroom.id,
                    student_id=student.id,
                    status="active"
                )
                db.add(enrollment)
                imported += 1
            else:
                skipped += 1

    except Exception as exc:
        errors.append(f"CSV parsing error: {str(exc)}")

    db.commit()
    return imported, skipped, errors


def export_classroom_to_excel(classroom: Classroom, db: Session) -> bytes:
    """
    Export classroom roster with attendance data to Excel file.

    Columns: Name | Phone | Email | External Ref | Tax ID | Gender | Status | Lessons Attended | Total Lessons
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Κατάλογος Τμήματος"

    headers = [
        "Ονοματεπώνυμο",
        "Τηλέφωνο",
        "Email",
        "Εξ. Ref",
        "ΑΦΜ",
        "Φύλο",
        "Κατάσταση",
        "Μαθήματα",
        "Παρουσίες"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    enrollments = db.query(Enrollment).filter(Enrollment.classroom_id == classroom.id).all()

    for row_num, enrollment in enumerate(enrollments, 2):
        student = enrollment.student
        attendances = db.query(Attendance).filter(
            Attendance.student_id == student.id,
            Attendance.lesson_id.in_(
                db.query(Lesson.id).filter(Lesson.classroom_id == classroom.id)
            )
        ).all()

        total_lessons = len(attendances)
        present_count = sum(1 for a in attendances if a.status == "present")

        row_data = [
            student.full_name,
            student.phone,
            student.email,
            student.external_ref,
            student.tax_id,
            student.gender,
            enrollment.status,
            total_lessons,
            present_count
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in (8, 9):  # Numbers
                cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_csv_template() -> str:
    """Generate a CSV template for classroom import."""
    return """full_name,phone,email,external_ref,tax_id,gender
Γιάννης Παπαδόπουλος,+306901234567,giannis@example.com,EXT001,123456789,Αρσενικό
Μαρία Κωνσταντινίδου,+306902345678,maria@example.com,EXT002,987654321,Θηλυκό
"""
